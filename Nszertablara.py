import tkinter as tk
import openpyxl
import json
from tkinter import font as tkFont
import os

window = tk.Tk()
window.title("N.szerelés")
window.attributes('-fullscreen', True)
window.columnconfigure(0, weight=1)
window.rowconfigure(0, weight=1)
window.rowconfigure(1, weight=1)
window.rowconfigure(2, weight=1)
window.rowconfigure(3, weight=1)
window.rowconfigure(4, weight=1)

frame1 = tk.Frame(window)
frame2 = tk.Frame(window)
frame2.config(bg="#ffffff")


def load_config(config_file):
    with open(config_file, 'r') as file:
        config = json.load(file)
    return config


config = load_config('config.json')
file_paths = config['file_paths']
data = file_paths['data']
Nszer = file_paths['nszer']
background = file_paths['Nszerbackground']

frame1.columnconfigure(0, weight=1)
for i in range(4):
    frame1.rowconfigure(i, weight=1)

frame1.grid(row=0, column=0, rowspan=5, sticky='nsew')
frame2.grid(row=4, column=0, sticky='nsew')
canvas = tk.Canvas(frame1)
canvas.grid(row=0, column=0, rowspan=4, sticky='nsew')
image = tk.PhotoImage(file=background)
canvas.create_image(0, 0, anchor=tk.NW, image=image)


x_scroll = tk.Scrollbar(frame1, orient='horizontal', command=canvas.xview)
x_scroll.grid(row=4, column=0, sticky='ew')
y_scroll = tk.Scrollbar(frame1, orient='vertical', command=canvas.yview)
y_scroll.grid(row=0, column=1, sticky='ns')
canvas.configure(xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)
canvas.configure(scrollregion=canvas.bbox('all'))

imported_widgets = {}

delete_widget_id = set()


def save_positions():
    file_path = Nszer
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            widget_positions = json.load(f)
    else:
        widget_positions = {}

    for widget in frame1.winfo_children():
        if isinstance(widget, tk.Text):
            widget_id = widget.widget_id
            color = widget.cget('bg')
            try:
                fg_color2 = widget.tag_cget("second_row", "foreground")
                fg_color3 = widget.tag_cget("third_row", "foreground")
                fg_color4 = widget.tag_cget("fourth_row", "foreground")
                fg_color5 = widget.tag_cget("fifth_row", "foreground")
            except tk.TclError:
                # If the tag doesn't exist, default to None or a fallback color
                fg_color2 = fg_color3 = fg_color4 = fg_color5 = "#000000"
            if widget_id in widget_positions:
                widget_positions[widget_id].update({
                    'color': color,
                    'second_row_color': fg_color2,
                    'third_row_color': fg_color3,
                    'fourth_row_color': fg_color4,
                    'fifth_row_color': fg_color5
                })
            else:
                widget_positions[widget_id] = {
                    "color": color,
                    "second_row_color": fg_color2,
                    "third_row_color": fg_color3,
                    "fourth_row_color": fg_color4,
                    "fifth_row_color": fg_color5
                }

    with open(file_path, 'w') as f:
        json.dump(widget_positions, f, indent=5)

    print("Widget color positions saved successfully!")


def restore_positions():
    try:
        with open(Nszer, 'r') as f:
            widget_positions = json.load(f)
            print("Widget positions restored successfully!")
    except FileNotFoundError:
        print("No widget positions found!")
        return

    for widget in frame2.winfo_children():
        if isinstance(widget, tk.Text) and widget.widget_id in widget_positions:
            widget.destroy()

    window.update()
    window.after(100)

    for widget in frame1.winfo_children():
        if isinstance(widget, tk.Text) and widget.widget_id not in widget_positions:
            widget.destroy()

    for widget_id, position in widget_positions.items():
        widget_data = imported_widgets.get(widget_id)
        color = position.get('color', '#FFFFFF')
        second_row_color = position.get('second_row_color', '#000000')
        third_row_color = position.get('third_row_color', '#000000')
        fourth_row_color = position.get('fourth_row_color', '#000000')
        fifth_row_color = position.get('fifth_row_color', '#000000')

        text_widget = None
        if widget_data:
            widget_exists_frame1 = False
            for frame1_widget in frame1.winfo_children():
                if isinstance(frame1_widget, tk.Text) and frame1_widget.widget_id == widget_id:
                    widget_exists_frame1 = True
                    canvas.coords(frame1_widget.canvas_id, position['x'], position['y'])
                    frame1_widget.config(bg=color)
                    frame1_widget.tag_configure("second_row", foreground=second_row_color)
                    frame1_widget.tag_add("second_row", "2.0", "2.end")
                    frame1_widget.tag_configure("third_row", foreground=third_row_color)
                    frame1_widget.tag_add("third_row", "3.0", "3.end")
                    frame1_widget.tag_configure("fourth_row", foreground=fourth_row_color)
                    frame1_widget.tag_add("fourth_row", "4.0", "4.end")
                    frame1_widget.tag_configure("fifth_row", foreground=fifth_row_color)
                    frame1_widget.tag_add("fifth_row", "5.0", "5.end")
                    break

            if not widget_exists_frame1:
                text_widget = tk.Text(frame1, wrap=tk.WORD, height=6, bd=2, borderwidth=2, relief='groove')
                text_widget.lift()
                text_widget.insert(tk.END, widget_data["text"])
                text_widget.config(width=widget_data["width"], bg=color)
                text_widget.tag_configure("center", justify='center')
                text_widget.tag_add("center", "1.0", "end")
                text_widget.tag_configure("second_row", foreground=second_row_color)
                text_widget.tag_add("second_row", "2.0", "2.end")
                text_widget.tag_configure("third_row", foreground=third_row_color)
                text_widget.tag_add("third_row", "3.0", "3.end")
                text_widget.tag_configure("fourth_row", foreground=fourth_row_color)
                text_widget.tag_add("fourth_row", "4.0", "4.end")
                text_widget.tag_configure("fifth_row", foreground=fifth_row_color)
                text_widget.tag_add("fifth_row", "5.0", "5.end")
                text_widget.grid(sticky='w')
                text_widget.config(state=tk.DISABLED)

                x = position['x']
                y = position['y']
                text_widget.bind("<Button-1>", lambda event: "break")
                canvas_id = canvas.create_window(x, y, anchor=tk.NW, window=text_widget)
                text_widget.canvas_id = canvas_id
                text_widget.bind("<Button-3>", show_right_click_menu)
                text_widget.widget_id = widget_id
    window.after(5000, restore_positions)
    recreate_widgets2()


def recreate_widgets2():
    try:
        with open(Nszer) as f:
            excluded_widget_positions = json.load(f)
    except FileNotFoundError:
        excluded_widget_positions = {}

    for widget in frame2.winfo_children():
        if not isinstance(widget, tk.Button):
            widget.destroy()

    workbook = openpyxl.load_workbook(data)
    sheet = workbook["N_szer"]

    num_columns = sheet.max_column
    column_data = []
    x_position = 0
    widget_height = 6
    for col_index in range(1, num_columns + 1):
        column_values = []
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            column_values.append(cell_value)
        if len(set(column_values)) == 1:
            continue
        column_data.append(column_values)

    unique_text_values = set()
    for column_values in column_data:
        column_tuple = tuple(column_values)
        if column_tuple in unique_text_values:
            continue
        unique_text_values.add(column_tuple)
        widget_id = str(column_values[0])
        if widget_id in excluded_widget_positions:
            continue
        imported_widgets[widget_id] = {
            "text": '\n'.join(str(value) for value in column_values[:6]),
            "width": int(column_values[7] * 10),
        }
        text_widget = tk.Text(frame2, wrap=tk.WORD, height=5, borderwidth=2, relief='groove')
        text_widget.lift()
        width = imported_widgets[widget_id]["width"]
        text_widget.insert(tk.END, imported_widgets[widget_id]["text"])
        if 5 <= width < 10:
            font_size = 7
        else:
            font_size = 9
        text_widget.tag_configure("center", justify='center', font=(None, font_size))
        text_widget.tag_add("center", "1.0", "end")
        text_widget.config(width=width, height=widget_height)
        text_widget.place(x=x_position, y=30)
        widget_width = imported_widgets[widget_id]["width"]
        x_position += widget_width * 8
        text_widget.bind("<Button-3>", show_right_click_menu)
        text_widget.widget_id = widget_id
    print("Recreated except with json excluded.")


def import_data():
    workbook = openpyxl.load_workbook(data, read_only= True)
    sheet = workbook["N_szer"]

    num_columns = sheet.max_column
    column_data = []
    x_position = 0
    widget_height = 6
    for col_index in range(1, num_columns + 1):
        column_values = []
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            column_values.append(cell_value)
        if len(set(column_values)) == 1:
            continue
        column_data.append(column_values)

    unique_text_values = set()
    for column_values in column_data:
        column_tuple = tuple(column_values)
        if column_tuple in unique_text_values:
            continue
        unique_text_values.add(column_tuple)
        widget_id = str(column_values[0])
        if widget_id in imported_widgets:
            continue
        imported_widgets[widget_id] = {
            "text": '\n'.join(str(value) for value in column_values[:6]),
            "width": int(column_values[7] * 10),
        }
        text_widget = tk.Text(frame2, wrap=tk.WORD, height=5, borderwidth=2, relief='groove')
        text_widget.lift()
        width = imported_widgets[widget_id]["width"]
        text_widget.insert(tk.END, imported_widgets[widget_id]["text"])
        if 5 <= width < 10:
            font_size = 7
        else:
            font_size = 9
        text_widget.tag_configure("center", justify='center', font=(None, font_size))
        text_widget.tag_add("center", "1.0", "end")
        text_widget.config(width=width, height=widget_height)
        text_widget.place(x=x_position, y=30)
        widget_width = imported_widgets[widget_id]["width"]
        x_position += widget_width * 8
        text_widget.bind("<Button-3>", show_right_click_menu)
        text_widget.widget_id = widget_id

    print("Data imported and widgets created.")


def move_widget_to_frame2(event):
    widget = event.widget
    widget_id = widget.widget_id

    for frame2_widget in frame2.winfo_children():
        if isinstance(frame2_widget, tk.Text) and frame2_widget.widget_id == widget_id:
            frame2_widget.destroy()
            break

    if widget_id in imported_widgets:
        del imported_widgets[widget_id]

    try:
        with open(Nszer, 'r') as f:
            widget_positions = json.load(f)
            if widget_id in widget_positions:
                del widget_positions[widget_id]
                print("Widget position deleted from the JSON file.")
        with open(Nszer, 'w') as f:
            json.dump(widget_positions, f)
    except FileNotFoundError:
        print("No widget positions found!")

    widget.destroy()

    print("Widget deleted successfully.")
    import_data()
    recreate_widgets2()


def show_right_click_menu(event):

    right_click_menu = tk.Menu(window, tearoff=0, bg='white', activeforeground='white',
                               activebackground='#8c43c4', fg='black')
    font_options = {'font': ('Helvetica', 10)}
    bold_font_options = {'font': tkFont.Font(family='Helvetica', size=10, weight='bold')}
    right_click_menu.add_command(label="1 oszlop kész", command=lambda: egy(event), **font_options)
    right_click_menu.add_separator()
    right_click_menu.add_command(label="2 oszlop kész", command=lambda: ketto(event), **font_options)
    right_click_menu.add_separator()
    right_click_menu.add_command(label="3 oszlop kész", command=lambda: tetel_kesz(event), **font_options)
    right_click_menu.add_separator()
    right_click_menu.add_command(label="second row", command=lambda: second_row(event), **bold_font_options)
    right_click_menu.add_separator()
    right_click_menu.add_command(label="third row", command=lambda: third_row(event), **bold_font_options)
    right_click_menu.add_separator()
    right_click_menu.add_command(label="fourth row", command=lambda: fourth_row(event), **bold_font_options)
    right_click_menu.add_separator()
    right_click_menu.add_command(label="fifth row", command=lambda: fifth_row(event), **bold_font_options)

    try:
        right_click_menu.tk_popup(event.x_root, event.y_root, 0)
    finally:
        right_click_menu.grab_release()
        return "break"


def adjust_widget_positions(delta_x):
    for widget in frame2.winfo_children():
        if isinstance(widget, tk.Text):
            current_x = widget.winfo_x()
            widget.place(x=current_x + delta_x)


def tetel_kesz(event):
    widget = event.widget
    user_color = '#e97117'

    try:
        widget.config(bg=user_color)
        widget.current_color = user_color
    except tk.TclError as e:
         print(f"Not correct color: {user_color}")
    save_positions()

def ketto(event):
    widget = event.widget
    user_color = '#e1a33c'

    try:
        widget.config(bg=user_color)
        widget.current_color = user_color
    except tk.TclError as e:
         print(f"Not correct color: {user_color}")
    save_positions()

def egy(event):
    widget = event.widget
    user_color = '#e9db7c'

    try:
        widget.config(bg=user_color)
        widget.current_color = user_color
    except tk.TclError as e:
         print(f"Not correct color: {user_color}")
    save_positions()


def second_row(event):
    widget = event.widget
    red_color = "#6343f9"
    black_color = "#000000"
    start = "2.0"
    end = "2.end"
    try:
        # Check the current color of the second row
        current_color = widget.tag_cget("second_row", "foreground")

        # Toggle the color based on current color
        if current_color == red_color:
            new_color = black_color
        else:
            new_color = red_color

        # Remove the existing tag
        widget.tag_remove("second_row", start, end)

        widget.tag_add("second_row", start, end)
        widget.tag_config("second_row", foreground=new_color)
        print(f"Second row color set to: {new_color}")

    except tk.TclError:
        # If the tag doesn't exist, create it with the red color
        widget.tag_add("second_row", start, end)
        widget.tag_config("second_row", foreground=red_color)
        print(f"Tag not found. Defaulting color to: {red_color}")
    save_positions()


def third_row(event):
    widget = event.widget
    green_color = "#6343f9"
    black_color = "#000000"
    start = "3.0"
    end = "3.end"
    try:
        # Check the current color of the second row
        current_color = widget.tag_cget("third_row", "foreground")

        # Toggle the color based on current color
        if current_color == green_color:
            new_color = black_color
        else:
            new_color = green_color

        # Remove the existing tag
        widget.tag_remove("third_row", start, end)

        widget.tag_add("third_row", start, end)
        widget.tag_config("third_row", foreground=new_color)
        print(f"Second row color set to: {new_color}")

    except tk.TclError:
        # If the tag doesn't exist, create it with the red color
        widget.tag_add("third_row", start, end)
        widget.tag_config("third_row", foreground=green_color)
        print(f"Tag not found. Defaulting color to: {green_color}")
    save_positions()


def fourth_row(event):
    widget = event.widget
    blue_color = "#6343f9"
    black_color = "#000000"
    start = "4.0"
    end = "4.end"
    try:
        # Check the current color of the second row
        current_color = widget.tag_cget("fourth_row", "foreground")

        # Toggle the color based on current color
        if current_color == blue_color:
            new_color = black_color
        else:
            new_color = blue_color

        # Remove the existing tag
        widget.tag_remove("fourth_row", start, end)

        widget.tag_add("fourth_row", start, end)
        widget.tag_config("fourth_row", foreground=new_color)
        print(f"Second row color set to: {new_color}")

    except tk.TclError:
        # If the tag doesn't exist, create it with the red color
        widget.tag_add("fourth_row", start, end)
        widget.tag_config("fourth_row", foreground=blue_color)
        print(f"Tag not found. Defaulting color to: {blue_color}")
    save_positions()


def fifth_row(event):
    widget = event.widget
    orange_color = "#6343f9"
    black_color = "#000000"
    start = "5.0"
    end = "5.end"
    try:
        # Check the current color of the second row
        current_color = widget.tag_cget("fifth_row", "foreground")

        # Toggle the color based on current color
        if current_color == orange_color:
            new_color = black_color
        else:
            new_color = orange_color

        # Remove the existing tag
        widget.tag_remove("fifth_row", start, end)

        widget.tag_add("fifth_row", start, end)
        widget.tag_config("fifth_row", foreground=new_color)
        print(f"Second row color set to: {new_color}")

    except tk.TclError:
        # If the tag doesn't exist, create it with the red color
        widget.tag_add("fifth_row", start, end)
        widget.tag_config("fifth_row", foreground=orange_color)
        print(f"Tag not found. Defaulting color to: {orange_color}")
    save_positions()


import_button = tk.Button(frame2, text="Import Data", command=import_data, bg="white",
                          activebackground="#ededed", activeforeground="black")
import_button.place(relx=0.98, anchor="ne")

save_button = tk.Button(frame2, text="Save Positions", command=save_positions, bg="white", activebackground="#ededed")
save_button.place(relx=0.93, anchor="ne")

restore_button = tk.Button(frame2, text="Restore Positions", command=restore_positions, bg="white",
                           activebackground="#ededed")
restore_button.place(relx=0.872, anchor="ne")

left_button = tk.Button(frame2, text="<-", command=lambda: adjust_widget_positions(20), bg="white",
                        activebackground="#ededed")
left_button.place(relx=0.785, anchor="ne")

right_button = tk.Button(frame2, text="->", command=lambda: adjust_widget_positions(-20), bg="white",
                         activebackground="#ededed")
right_button.place(relx=0.80, anchor="ne")

frame2.grid_remove()

import_data()
restore_positions()

window.mainloop()
