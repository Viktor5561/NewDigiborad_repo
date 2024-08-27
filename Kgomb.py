import tkinter as tk
import json
import openpyxl
import os
from tkinter import font as tkFont

window = tk.Tk()
window.title("N.gombolyítás")
window.state('zoomed')


window.columnconfigure(0, weight=1)
window.rowconfigure(0, weight=1)
window.rowconfigure(1, weight=1)
window.rowconfigure(2, weight=1)
window.rowconfigure(3, weight=1)
window.rowconfigure(4, weight=2)

frame1 = tk.Frame(window)
frame2 = tk.Frame(window)
frame2.config(bg="#ffffff")


def load_config(config_file):
    with open(config_file, 'r') as file:
        config = json.load(file)
    return config


config = load_config('config.json')
file_paths = config['file_paths']
data = file_paths['data']
Kgomb = file_paths['ngomb']
#background = file_paths['Kgombbackground']

frame1.columnconfigure(0, weight=1)
for i in range(4):
    frame1.rowconfigure(i, weight=1)

frame1.grid(row=0, column=0, rowspan=4, sticky='nsew')
frame2.grid(row=4, column=0, sticky='nsew')
canvas = tk.Canvas(frame1)
canvas.grid(row=0, column=0, rowspan=4, sticky='nsew')
image = tk.PhotoImage(file="Ngomb_bground.png")
canvas.create_image(0, 0, anchor=tk.NW, image=image)


x_scroll = tk.Scrollbar(frame1, orient='horizontal', command=canvas.xview)
x_scroll.grid(row=4, column=0, sticky='ew')
y_scroll = tk.Scrollbar(frame1, orient='vertical', command=canvas.yview)
y_scroll.grid(row=0, column=1, sticky='ns')
canvas.configure(xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)
canvas.configure(scrollregion=canvas.bbox('all'))

imported_widgets = {}

delete_widget_id = set()


def start_drag(event):
    widget = event.widget
    widget.startX = event.x
    widget.startY = event.y
    widget.configure(cursor="hand1")
    widget.configure(insertofftime=0)
    widget.lift()

    widget.startXPos = event.x_root - widget.winfo_rootx()
    widget.startYPos = event.y_root - widget.winfo_rooty()


def drag_widget(event):
    widget = event.widget
    scroll_x = canvas.canvasx(0)
    scroll_y = canvas.canvasy(0)
    x = event.x_root - widget.startXPos - window.winfo_rootx() + scroll_x
    y = event.y_root - widget.startYPos - window.winfo_rooty() + scroll_y
    canvas.coords(widget.canvas_id, x, y)


def stop_drag(event):
    event.widget.configure(cursor="")
    event.widget.configure(insertofftime=600)


def widget_to_frame1(event):
    global delete_widget_id

    widget = event.widget
    widget_id = widget.widget_id

    for frame2_widget in frame2.winfo_children():
        if isinstance(frame2_widget, tk.Text) and getattr(frame2_widget, 'widget_id', None) == widget_id:
            frame2_widget.destroy()
            break

    delete_widget_id.add(widget_id)

    widget_exists_in_frame1 = False
    for frame1_widget in frame1.winfo_children():
        if getattr(frame1_widget, 'widget_id', None) == widget_id:
            widget_exists_in_frame1 = True
            break

    if widget_id in imported_widgets and not widget_exists_in_frame1:
        widget_data = imported_widgets[widget_id]

        widget_height = 6
        text_widget = tk.Text(frame1, wrap=tk.WORD, height=5, bd=2, borderwidth=2, relief='groove')
        text_widget.insert(tk.END, widget_data["text"])
        text_widget.config(width=widget_data["width"], height=widget_height)
        text_widget.tag_configure("center", justify='center')
        text_widget.tag_add("center", "1.0", "end")
        text_widget.grid(sticky='w')

        x = 50
        y = 50
        canvas_id = canvas.create_window(x, y, anchor=tk.NW, window=text_widget)
        text_widget.canvas_id = canvas_id
        text_widget.bind("<Button-1>", start_drag)
        text_widget.bind("<B1-Motion>", drag_widget)
        text_widget.bind("<ButtonRelease-1>", stop_drag)
        text_widget.bind("<Button-3>", show_right_click_menu)
        text_widget.bind("<ButtonPress-1>", start_drag)
        text_widget.widget_id = widget_id
        text_widget.configure(state="disabled")

    recreate_widgets()
    print("Widget deleted successfully.")


def recreate_widgets():
    global delete_widget_id

    with open(Kgomb, 'r') as f:
        saved_widget_positions = json.load(f)
    saved_widget_ids = set(saved_widget_positions.keys())

    for widget in frame2.winfo_children():
        if not isinstance(widget, tk.Button):
            widget.destroy()

    workbook = openpyxl.load_workbook(data, read_only= True)
    sheet = workbook["HV"]

    num_columns = sheet.max_column
    column_data = []
    x_position = 0
    for col_index in range(1, num_columns + 1):
        column_values = []
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            column_values.append(cell_value)
        if len(set(column_values)) == 1:
            continue
        column_data.append(column_values)

    unique_text_values = set()
    for column_values in column_data:
        column_tuple = tuple(column_values)
        if column_tuple in unique_text_values:
            continue
        unique_text_values.add(column_tuple)
        widget_id = str(column_values[0])
        if widget_id in delete_widget_id or widget_id in saved_widget_ids:
            continue
        imported_widgets[widget_id] = {
            "text": '\n'.join(str(value) for value in column_values[:6]),
            "width": int(column_values[7] * 10),
        }
        text_widget = tk.Text(frame2, wrap=tk.WORD, height=6, borderwidth=2, relief='groove')
        text_widget.lift()
        width = imported_widgets[widget_id]["width"]
        text_widget.insert(tk.END, imported_widgets[widget_id]["text"])
        if 5 <= width < 10:
            font_size = 7
        else:
            font_size = 9
        text_widget.tag_configure("center", justify='center', font=(None, font_size))
        text_widget.tag_add("center", "1.0", "end")
        text_widget.config(width=width)
        text_widget.place(x=x_position, y=30)
        widget_width = imported_widgets[widget_id]["width"]
        x_position += widget_width * 8
        text_widget.bind("<Button-3>", show_right_click_menu)
        text_widget.widget_id = widget_id
        text_widget.configure(state="disabled")
    print("Widgets recreated.")


def save_positions():
    widget_positions = {}

    file_path = Kgomb
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            current_data = json.load(f)
    else:
        current_data = {}

    for widget in frame1.winfo_children():
        if isinstance(widget, tk.Text):
            widget_id = widget.widget_id  # Retrieve the widget ID
            x, y = canvas.coords(widget.canvas_id)


            widget_data = current_data.get(widget_id, {})
            widget_data.update({"x": x, "y": y})
            widget_positions[widget_id] = widget_data

    with open(file_path, 'w') as f:
        json.dump(widget_positions, f, indent=5)

    print("Widget positions saved successfully!")


def restore_positions():
    try:
        with open(Kgomb, 'r') as f:
            widget_positions = json.load(f)
            print("Widget positions restored successfully!")
    except FileNotFoundError:
        print("No widget positions found!")
        return

    # Destroy the widgets in frame2
    for widget in frame2.winfo_children():
        if isinstance(widget, tk.Text) and widget.widget_id in widget_positions:
            widget.destroy()

    window.update()
    window.after(100)

    for widget_id, position in widget_positions.items():
        widget_data = imported_widgets.get(widget_id)
        color = position.get('color', '#FFFFFF')
        second_row_color = position.get('second_row_color', '#000000')
        third_row_color = position.get('third_row_color', '#000000')
        fourth_row_color = position.get('fourth_row_color', '#000000')
        fifth_row_color = position.get('fifth_row_color', '#000000')

        text_widget = None
        if widget_data:
            widget_exists_frame1 = False
            for frame1_widget in frame1.winfo_children():
                if isinstance(frame1_widget, tk.Text) and frame1_widget.widget_id == widget_id:
                    widget_exists_frame1 = True
                    canvas.coords(frame1_widget.canvas_id, position['x'], position['y'])
                    frame1_widget.config(bg=color)
                    frame1_widget.tag_configure("second_row", foreground=second_row_color)
                    frame1_widget.tag_add("second_row", "2.0", "2.end")
                    frame1_widget.tag_configure("third_row", foreground=third_row_color)
                    frame1_widget.tag_add("third_row", "3.0", "3.end")
                    frame1_widget.tag_configure("fourth_row", foreground=fourth_row_color)
                    frame1_widget.tag_add("fourth_row", "4.0", "4.end")
                    frame1_widget.tag_configure("fifth_row", foreground=fifth_row_color)
                    frame1_widget.tag_add("fifth_row", "5.0", "5.end")
                    break

            if not widget_exists_frame1:
                text_widget = tk.Text(frame1, wrap=tk.WORD, height=6, bd=2, borderwidth=2, relief='groove')
                text_widget.lift()
                text_widget.insert(tk.END, widget_data["text"])
                text_widget.config(width=widget_data["width"], bg=color)
                text_widget.tag_configure("center", justify='center')
                text_widget.tag_add("center", "1.0", "end")
                text_widget.tag_configure("second_row", foreground=second_row_color)
                text_widget.tag_add("second_row", "2.0", "2.end")
                text_widget.tag_configure("third_row", foreground=third_row_color)
                text_widget.tag_add("third_row", "3.0", "3.end")
                text_widget.tag_configure("fourth_row", foreground=fourth_row_color)
                text_widget.tag_add("fourth_row", "4.0", "4.end")
                text_widget.tag_configure("fifth_row", foreground=fifth_row_color)
                text_widget.tag_add("fifth_row", "5.0", "5.end")
                text_widget.grid(sticky='w')
                text_widget.configure(state="disabled")

                x = position['x']
                y = position['y']
                canvas_id = canvas.create_window(x, y, anchor=tk.NW, window=text_widget)
                text_widget.canvas_id = canvas_id
                text_widget.bind("<Button-1>", start_drag)
                text_widget.bind("<B1-Motion>", drag_widget)
                text_widget.bind("<ButtonRelease-1>", stop_drag)
                text_widget.bind("<Button-3>", show_right_click_menu)
                text_widget.widget_id = widget_id


    update_frame2()


def update_frame2():
    try:
        with open(Kgomb) as f:
            excluded_widget_positions = json.load(f)
    except FileNotFoundError:
        excluded_widget_positions = {}

    for widget in frame2.winfo_children():
        if not isinstance(widget, tk.Button):
            widget.destroy()

    workbook = openpyxl.load_workbook(data, read_only= True)
    sheet = workbook["HV"]

    num_columns = sheet.max_column
    column_data = []
    x_position = 0
    widget_height = 6
    for col_index in range(1, num_columns + 1):
        column_values = []
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            column_values.append(cell_value)
        if len(set(column_values)) == 1:
            continue
        column_data.append(column_values)

    unique_text_values = set()
    for column_values in column_data:
        column_tuple = tuple(column_values)
        if column_tuple in unique_text_values:
            continue
        unique_text_values.add(column_tuple)
        widget_id = str(column_values[0])
        if widget_id in excluded_widget_positions:
            continue
        imported_widgets[widget_id] = {
            "text": '\n'.join(str(value) for value in column_values[:6]),
            "width": int(column_values[7] * 10),
        }
        text_widget = tk.Text(frame2, wrap=tk.WORD, height=widget_height, borderwidth=2, relief='groove')
        text_widget.lift()
        width = imported_widgets[widget_id]["width"]
        text_widget.insert(tk.END, imported_widgets[widget_id]["text"])
        if 5 <= width < 10:
            font_size = 7
        else:
            font_size = 9
        text_widget.tag_configure("center", justify='center', font=(None, font_size))
        text_widget.tag_add("center", "1.0", "end")
        text_widget.config(width=width)
        text_widget.place(x=x_position, y=30)
        text_widget.configure(state="disabled")
        widget_width = imported_widgets[widget_id]["width"]
        x_position += widget_width * 8
        text_widget.bind("<Button-3>", show_right_click_menu)
        text_widget.widget_id = widget_id
    print("Recreated except with json excluded.")


def import_data():
    workbook = openpyxl.load_workbook(data, read_only= True)
    sheet = workbook["HV"]

    num_columns = sheet.max_column
    column_data = []
    x_position = 0
    widget_height = 6
    for col_index in range(1, num_columns + 1):
        column_values = []
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            column_values.append(cell_value)
        if len(set(column_values)) == 1:
            continue
        column_data.append(column_values)

    unique_text_values = set()
    for column_values in column_data:
        column_tuple = tuple(column_values)
        if column_tuple in unique_text_values:
            continue
        unique_text_values.add(column_tuple)
        widget_id = str(column_values[0])
        if widget_id in imported_widgets:
            continue
        imported_widgets[widget_id] = {
            "text": '\n'.join(str(value) for value in column_values[:6]),
            "width": int(column_values[7] * 10),
        }
        text_widget = tk.Text(frame2, wrap=tk.WORD, height=5, borderwidth=2, relief='groove')
        text_widget.lift()
        width = imported_widgets[widget_id]["width"]
        text_widget.insert(tk.END, imported_widgets[widget_id]["text"])
        if 5 <= width < 10:
            font_size = 7
        else:
            font_size = 9
        text_widget.tag_configure("center", justify='center', font=(None, font_size))
        text_widget.tag_add("center", "1.0", "end")
        text_widget.config(width=width, height=widget_height)
        text_widget.place(x=x_position, y=30)
        widget_width = imported_widgets[widget_id]["width"]
        x_position += widget_width * 8
        text_widget.configure(state="disabled")
        text_widget.bind("<Button-3>", show_right_click_menu)
        text_widget.widget_id = widget_id
    print(imported_widgets)


def move_widget_to_frame2(event):
    widget = event.widget
    widget_id = widget.widget_id

    for frame2_widget in frame2.winfo_children():
        if isinstance(frame2_widget, tk.Text) and frame2_widget.widget_id == widget_id:
            frame2_widget.destroy()
            break

    if widget_id in imported_widgets:
        del imported_widgets[widget_id]

    try:
        with open(Kgomb, 'r') as f:
            widget_positions = json.load(f)
            if widget_id in widget_positions:
                del widget_positions[widget_id]
                print("Widget position deleted from the JSON file.")
        with open(Kgomb, 'w') as f:
            json.dump(widget_positions, f, indent=5)
    except FileNotFoundError:
        print("No widget positions found!")

    widget.destroy()

    print("Widget deleted successfully.")
    import_data()
    update_frame2()


def show_right_click_menu(event):

    right_click_menu = tk.Menu(window, tearoff=0, bg='white', activeforeground='white', activebackground='#8c43c4',
                               fg='black')
    font_options = {'font': ('Helvetica', 10)}
    bold_font_options = {'font': tkFont.Font(family='Helvetica', size=10, weight='bold')}
    right_click_menu.add_command(label="Táblára tesz", command=lambda: widget_to_frame1(event), **font_options)
    right_click_menu.add_separator()
    right_click_menu.add_command(label="Levétel tábláról", command=lambda: move_widget_to_frame2(event), **font_options)


    try:
        right_click_menu.tk_popup(event.x_root, event.y_root, 0)
    finally:
        right_click_menu.grab_release()
        return "break"


def adjust_widget_positions(delta_x):
    for widget in frame2.winfo_children():
        if isinstance(widget, tk.Text):
            current_x = widget.winfo_x()
            widget.place(x=current_x + delta_x)


import_button = tk.Button(frame2, text="Import Data", command=import_data, background="white",
                          activebackground="#ededed")
import_button.place(relx=0.98, anchor="ne")

save_button = tk.Button(frame2, text="Save Positions", command=save_positions, background="white",
                        activebackground="#ededed")
save_button.place(relx=0.93, anchor="ne")

restore_button = tk.Button(frame2, text="Restore Positions", command=restore_positions, background="white",
                           activebackground="#ededed")
restore_button.place(relx=0.872, anchor="ne")

left_button = tk.Button(frame2, text="<", command=lambda: adjust_widget_positions(20), background="white",
                        activebackground="#ededed")
left_button.place(relx=0.785, anchor="ne")

right_button = tk.Button(frame2, text=">", command=lambda: adjust_widget_positions(-20), background="white",
                         activebackground="#ededed")
right_button.place(relx=0.80, anchor="ne")

window.mainloop()
